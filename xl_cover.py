import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Color, Fill, Font, PatternFill, Border
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from settings import Settings


class XlCover():
    """
    Class that creates and formats a cover sheet
    """
    def __init__(self, wbook, sheet= None) -> None:
        self.settings = Settings().data["Cover Settings"]
        self.wbook = wbook
        self.sheet = sheet
        if sheet == None:
            self._create_cover_sheet()
        
    def _create_cover_sheet(self):
        self.sheet = self.wbook.create_sheet("Cover", index= 0)
        self._set_heads()
        self._set_titles()
        self._add_logo()
        self._set_info()
        self._signature_boxes()

    def _set_titles(self):
        for key, value in self.settings["Titles"].items():
            self._set_title(key, value)

    def _set_title(self, key, value):
        start = value["Start"]
        end = value["End"]
        self.sheet.merge_cells(f"{start}:{end}")
        self.sheet[start].value = key
        self.sheet[start].font = Font(color= value["Font Color"], size= value["Font Size"])
        self.sheet[start].fill = PatternFill(fgColor=value["Fill"], fill_type= 'solid')
        self.sheet[start].alignment = Alignment(horizontal= value["Alignment"], vertical= 'center', indent= value["Indents"])

    def _add_logo(self):
        logo = "scania-symbol.png"
        img = openpyxl.drawing.image.Image(logo)
        h, w = img.height, img.width
        p2e = pixels_to_EMU
        position = XDRPoint2D(p2e(520), p2e(10))
        size = XDRPositiveSize2D(p2e(w/2), p2e(h/2))
        anchor = AbsoluteAnchor(pos=position, ext=size)
        img.anchor = anchor
        self.sheet.add_image(img)

    def _set_heads(self):
        for key, value in self.settings["Heads"].items():
            self._set_head(key, value)

    def _set_head(self, key, value):
        cell = self.sheet[value["Start"]]
        if value["End"] is not None:
            self.sheet.merge_cells("{0}:{1}".format(value["Start"], value["End"]))
        cell.value = f"{key}:"
        cell.font = Font(size= value["Font Size"], bold= value["Bold"])
        cell.alignment = Alignment(horizontal= value["Alignment"])
        if value["Merged Input"] is not None:
            self.sheet.merge_cells(value["Merged Input"])
            self.sheet[value["Merged Input"].split(":")[0]].font = Font(size= value["Font Size Input"])

    def _set_info(self):
        info = self.settings["Info"]
        self.sheet.merge_cells(info["Merge Cells"])
        cell = self.sheet[info["Merge Cells"].split(":")[0]]
        cell.value = info["Text"]
        cell.font = Font(size= info["Font Size"])
        cell.alignment = Alignment(horizontal= info["Horizontal Alignment"], vertical= info["Vertical Alignment"], wrap_text= True)
        self._add_thick_outer_borders(info["Merge Cells"])

    def _add_thick_outer_borders(self, coordinates):
        start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(coordinates.split(":")[0])
        end_row, end_col = openpyxl.utils.cell.coordinate_to_tuple(coordinates.split(":")[1])
        thin = Side(style= 'thin')
        thick = Side(style= 'thick')
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                left, right, top, bottom = thin, thin, thin, thin
                if row == start_row:
                    top = thick
                if row == end_row:
                    bottom = thick
                if col == start_col:
                    left = thick
                if col == end_col:
                    right = thick
                self.sheet.cell(row= row, column= col).border = Border(left= left, right= right, top= top, bottom= bottom)

    def _signature_boxes(self):
        titles = self.settings["Box"]["Titles"]
        for title in titles:
            start = self.settings["Heads"][title]["Start"]
            self._signature_box(start)

    def _signature_box(self, start):
        start_row, start_column_index = openpyxl.utils.coordinate_to_tuple(start)
        start_column = openpyxl.utils.get_column_letter(start_column_index + 1)
        row = start_row
        for key, value in self.settings["Box"]["Heads"].items():
            start = start_column + str(row)
            end = openpyxl.utils.cell.get_column_letter(start_column_index + value["Head Width"]) + str(row + value["Head Height"] - 1)
            start_input = openpyxl.utils.cell.get_column_letter(start_column_index + value["Head Width"] + 1)  + str(row)
            end_input = openpyxl.utils.cell.get_column_letter(start_column_index + self.settings["Box"]["Total Width"])  + str(row + value["Head Height"] - 1)
            cell = self.sheet[start]
            cell.value = key + ":"
            cell.font = Font(bold= value["Bold"])
            if value["Head Width"] > 1 or value["Head Height"] > 1:
                self.sheet.merge_cells(f"{start}:{end}")
            self.sheet.merge_cells(f"{start_input}:{end_input}")
            row += value["Head Height"]
        end_column = openpyxl.utils.cell.get_column_letter(start_column_index + self.settings["Box"]["Total Width"])
        coordinates = f"{start_column}{start_row}:{end_column}{row - 1}"
        self._add_thick_outer_borders(coordinates= coordinates)

    def set_part_rev(self, partnum, rev):
        part_row, part_col = openpyxl.utils.coordinate_to_tuple(self.settings["Heads"]["Part Number"]["End"])
        rev_row, rev_col = openpyxl.utils.coordinate_to_tuple(self.settings["Heads"]["Revision"]["End"])
        self.sheet.cell(row= part_row , column= part_col + 1).value = partnum
        self.sheet.cell(row= rev_row, column= rev_col + 1).value = rev

if __name__ == "__main__":
    pass