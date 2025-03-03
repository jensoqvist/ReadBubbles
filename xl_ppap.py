import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Color, Fill, Font, PatternFill, Border
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Series, Reference


class XlPpap():
    def __init__(self, settings, wbook, xlhandler, df, org_tablename, org_table_pos) -> None:
        self.settings = settings.data
        self.wbook = wbook
        self.xl = xlhandler
        self.df = df
        self.org_tablename = org_tablename
        self.org_table_pos = org_table_pos
        self.sheet_name = xlhandler.sheet_name + " PPAP"
        self.sheet = None
        self.columns = self.settings["PPAP"]["Columns"]
        self.table_start_col_index = 1 + self.xl.start_col
        self.table_start_row_index = self.xl.skip_rows + 1
        self.tablestart = f"{openpyxl.utils.get_column_letter(self.table_start_col_index)}{self.xl.skip_rows + 1}"
        self.table_end_row_index = len(self.df) + self.xl.skip_rows + 1
        self.col_len = len(self.settings["PPAP"]["Columns"])
        self.table_end_col_index = self.col_len + self.xl.start_col
        self.table_end = f"{openpyxl.utils.get_column_letter(self.table_end_col_index)}{self.table_end_row_index}"
        self.tablename = f"Positions_REV_{self.xl.revnum}_PPAP"
        self.start()

    def start(self):
        self._create_ppap_sheet()
        self._write_positions()
        self._add_column_names()
        self._add_table(self.tablename, self.tablestart, self.table_end)
        self._table_colors()
        self._add_headings()
        self._adjust_sizes()
        self._copy_columns()
        self._alignment()
        self._add_borders(start_row= self.table_start_row_index, end_row= self.sheet.max_row + 1, start_col= self.table_start_col_index, end_col= self.table_end_col_index + 1) # Borders in Table
        self._data_validation()
        self._conditional_formating()
        self._set_condition_amount()
        self._set_condition_ppk()
        self._add_limits()
        self._create_barchart()
        self._add_count()
        self._add_text()

    def _create_ppap_sheet(self):
        if self.sheet_name not in self.wbook.sheetnames:
            self.sheet = self.wbook.create_sheet(self.sheet_name)   
        else:
            self.sheet = self.wbook[self.sheet_name] 

    def _write_positions(self):
        col = openpyxl.utils.get_column_letter(self.table_start_col_index)
        for index, pos in enumerate(self.df):
            row = int(self.table_start_row_index) + index + 1
            self.sheet[f"{col}{row}"] = pos

    def _add_column_names(self):
        for index, col in enumerate(self.columns):
            cell = self.sheet[openpyxl.utils.get_column_letter(self.xl.start_col + index + 1) + str(self.xl.skip_rows + 1)]
            cell.value = col
            cell.alignment = Alignment(horizontal='center', vertical= 'top')

    def _add_table(self, tablename, start, end):
        self.table = Table(displayName= tablename, ref=f'{start}:{end}')
        style = TableStyleInfo(name="TableStyleMedium16", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        self.table.tableStyleInfo = style
        self.sheet.add_table(self.table)

    def _table_colors(self):
        color = self.settings["Colors"]["Scania Blue"]
        for col in range(self.table_start_col_index, self.table_end_col_index + 1):
            cell = self.sheet[f"{openpyxl.utils.get_column_letter(col)}{self.table_start_row_index}"]
            cell.font = Font(color= color["Font Color"])
            cell.fill = PatternFill(fgColor= color["Fill Color"], fill_type= 'solid')


    def _add_headings(self):
        self.sheet['B2'].value = "PART NUMBER"
        self.sheet['B2'].font = Font(bold= True)
        self.sheet['C2'].value = self.xl.partnum
        self.sheet['B3'].value = "REVISION"
        self.sheet['B3'].font = Font(bold= True)
        self.sheet['C3'].value = self.xl.revnum
        self._add_thick_outer_borders(start_row = 2, end_row= 3, start_col = 2, end_col= 3)

    def _adjust_sizes(self):   
        self.sheet.column_dimensions["A"].width = 4
        self.sheet.row_dimensions[self.table_start_row_index].height = 30
        for i, col in enumerate(self.columns):
            col_index = i +  self.table_start_col_index
            try:
                self.sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = self.settings["Column Size"][col]
            except:
                self.sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = 4


    def _alignment(self):
        for r in range(self.table_start_row_index + 1, self.sheet.max_row + 1):
            self.sheet.cell(r, self.table_start_col_index).alignment = Alignment(horizontal='right')


    def _add_borders(self, start_row, end_row, start_col, end_col):
        side = Side(style= 'thin')
        border = Border(left= side, right= side, top= side, bottom= side)
        for row in range(start_row, end_row):
            for col in range(start_col, end_col):
                self.sheet.cell(row= row, column= col).border = border         

    def _add_thick_outer_borders(self, start_row, end_row, start_col, end_col):
        thin = Side(style= 'thin')
        thick = Side(style= 'thick')
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                left, right, top, bottom = thin, thin, thin, thin
                if row == start_row:
                    top = thick
                if row == end_row:
                    bottom = thick
                if col == start_col:
                    left = thick
                if col == end_col:
                    right = thick
                self.sheet.cell(row= row, column= col).border = Border(left= left, right= right, top= top, bottom= bottom)

    def _data_validation(self):
        dic = self.settings["Data Validation"]
        dic.update(self.settings["PPAP"]["Data Validation"])
        for key, value in dic.items():
            validation_string = ", ".join(value)
            if key in self.columns:
                col_index = self.columns.index(key) + self.table_start_col_index
                dv = DataValidation(type= "list", formula1= f'"{validation_string}"', allow_blank= True)
                dv.error = "Invalid Entry"
                dv.errorTitle = "Invalid Entry"
                self.sheet.add_data_validation(dv)
                for r in range(self.table_start_row_index + 1, self.sheet.max_row + 1):
                    dv.add(openpyxl.utils.get_column_letter(col_index) + str(r)) 

    def _conditional_formating(self):
        dic = self.settings["Conditional"]
        dic.update(self.settings["PPAP"]["Conditional"])
        for colname, words in dic.items():
            self._conditional_col(colname, words)

    def _conditional_col(self, colname, words):  
        rownum = self.table_start_row_index  - 1
        for col in self.sheet.iter_cols(self.table_start_col_index, self.sheet.max_column):
            col_letter = openpyxl.utils.get_column_letter(col[rownum].column)
            if col[rownum].value == colname:
                for word, color in words.items():
                    colors = self.settings["Colors"][color]
                    fill = PatternFill()
                    if colors["Fill Color"] is not None:
                        fill = PatternFill( bgColor= colors["Fill Color"], fill_type='solid')
                    self._set_condition(col_letter, Font(color= colors["Font Color"]), fill, word)

    def _set_condition(self, col, text, fill, word):
        dxf = DifferentialStyle(font= text, fill=fill)
        rule = Rule(type="containsText", operator="containsText", text= word, dxf=dxf)
        rule.formula = [f'NOT(ISERROR(SEARCH("{word}",{col}{self.table_start_row_index})))']
        self.sheet.conditional_formatting.add(f'{col}{self.table_start_row_index}:{col}{self.table_end_row_index}', rule)

    def _set_condition_amount(self):
        col = openpyxl.utils.get_column_letter(self.table_start_col_index + self.settings["PPAP"]["Columns"].index("Amount Measured"))
        color = self.settings["Colors"]["Red"]
        fill = PatternFill( bgColor= color["Fill Color"], fill_type='solid')
        rule= CellIsRule(operator='lessThan', formula=[self.settings["PPAP"]["Conditional"]["Conditional Amount"]], stopIfTrue=True, fill= fill)
        self.sheet.conditional_formatting.add(f'{col}{self.table_start_row_index}:{col}{self.table_end_row_index}', rule)

    def _set_condition_ppk(self):
        col = openpyxl.utils.get_column_letter(self.table_start_col_index + self.settings["PPAP"]["Columns"].index("Ppk/Cpk"))
        points = self.settings["PPAP"]["Conditional"]["Conditional Ppk"]
        rule= ColorScaleRule(start_type='num', start_value=points["start"], start_color=self.settings["Colors"]["Red"]["Fill Color"],
                           mid_type='num', mid_value=points["mid"], mid_color=self.settings["Colors"]["Orange"]["Fill Color"],
                           end_type='num', end_value=points["end"], end_color=self.settings["Colors"]["Green"]["Fill Color"])
        self.sheet.conditional_formatting.add(f'{col}{self.table_start_row_index}:{col}{self.table_end_row_index}', rule)

    def _copy_columns(self):
        copy_columns = self.settings["PPAP"]["Copy Columns"]
        org_columns = ["Type"] + [*self.settings["Columns"].keys()]
        for index, col in enumerate(self.columns):
            if col in copy_columns:
                org_index = org_columns.index(col)
                self._copy_row(index, org_index)

    def _copy_row(self, index, org_index):
        col = openpyxl.utils.get_column_letter(self.xl.start_col + index + 1)
        for row in range(self.table_start_row_index + 1, self.table_end_row_index + 1):
            vlook = f'=IF(VLOOKUP($B{str(row)},{self.org_table_pos},{org_index + 2},FALSE)=0,"",VLOOKUP($B{str(row)},{self.org_table_pos},{org_index + 2},FALSE))'
            self.sheet[f"{col}{row}"] = vlook

    def _add_limits(self):
        col = openpyxl.utils.get_column_letter(self.table_start_col_index + self.settings["PPAP"]["Columns"].index("Limit"))
        col_clas = openpyxl.utils.get_column_letter(self.table_start_col_index + self.settings["PPAP"]["Columns"].index("Classification"))
        for row in range(self.table_start_row_index + 1, self.table_end_row_index + 1):
            formula = f'=IF({col_clas}{row} = "<S>";1;IF({col_clas}{row} = "<M>";1.33;IF({col_clas}{row} = "";0;1.67)))'.replace(";", ",")
            self.sheet[f"{col}{row}"] = formula
        
    def _create_barchart(self):
        #Chart settings
        chart = BarChart()
        chart.type = self.settings["PPAP"]["Chart"]["Type"]
        chart.style = self.settings["PPAP"]["Chart"]["Style"]
        chart.overlap = self.settings["PPAP"]["Chart"]["Overlapp"]
        chart.gapWidth = self.settings["PPAP"]["Chart"]["Gap Width"]
        chart.x_axis.scaling.orientation = self.settings["PPAP"]["Chart"]["Orientation"]
        chart.height = (14.4 * ((self.table_end_row_index - self.table_start_row_index + 1)))  / 28.4
        chart.width = self.settings["PPAP"]["Chart"]["Width"]
        chart.x_axis.delete = False
        chart.y_axis.delete = False

        #Data
        data_col = openpyxl.utils.get_column_letter(self.table_start_col_index + self.settings["PPAP"]["Columns"].index("Ppk/Cpk"))
        data_col_limit = openpyxl.utils.get_column_letter(self.table_start_col_index + self.settings["PPAP"]["Columns"].index("Limit"))
        data_ppk = Reference(worksheet= self.sheet, range_string= f"'{self.sheet_name}'!${data_col}${self.table_start_row_index + 1}:${data_col}${self.table_end_row_index}")
        data_limit = Reference(worksheet= self.sheet, range_string= f"'{self.sheet_name}'!${data_col_limit}${self.table_start_row_index + 1}:${data_col_limit}${self.table_end_row_index}")

        #Series settings
        series_ppk = Series(data_ppk, title= "Ppk/Cpk")
        series_ppk.graphicalProperties.solidFill = self.settings["PPAP"]["Chart"]["Ppk Color"]
        series_ppk.graphicalProperties.line.noFill = True
        series_limit = Series(data_limit, title= "Limit")
        series_limit.graphicalProperties.solidFill = self.settings["PPAP"]["Chart"]["Limit Color"]
        series_limit.graphicalProperties.line.solidFill = self.settings["PPAP"]["Chart"]["Limit Color"]
        series_limit.graphicalProperties.line.dashStyle = "sysDot"
        cats_col  = openpyxl.utils.get_column_letter(self.table_start_col_index + self.settings["PPAP"]["Columns"].index("Position Number"))
        cats = Reference(worksheet= self.sheet, range_string= f"'{self.sheet_name}'!${cats_col}${self.table_start_row_index + 1}:${cats_col}${self.table_end_row_index}")

        #Setup
        chart.append(series_limit)
        chart.append(series_ppk)
        chart.set_categories(cats)
        self.sheet.add_chart(chart, "M6")


    def _add_count(self):
        self.sheet['D2'].value = "Action Count:"
        self.sheet['D2'].font = Font(bold= True)
        self.sheet['D3'].value = f'=COUNTIF({self.tablename}[Action], "Yes")'
        self._add_thick_outer_borders(start_row = 2, end_row= 3, start_col = 4, end_col= 8)


    def _add_text(self):
        text = "This sheet is intended for visualization of the PPAP process and decided actions. Responsible: DXT"
        self.sheet['F2'].value = text
        self.sheet['F2'].font = Font(bold= True)
