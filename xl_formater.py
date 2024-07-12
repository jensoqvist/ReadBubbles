import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Color, Fill, Font, PatternFill, Border
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from settings import Settings
from xl_cover import XlCover

class XlFormater():
    """
    Class to give .xlsx file the correct format after the dataframe has been added to it.\n\n

    Parameters:\n
        xlhandler\n 
        df_handler\n
        duplicates\n\n

    Attributes:\n
        self.settings = Settings()\n
        self.xl = xlhandler\n
        self.df_handler = df_handler\n
        self.df = df_handler.df\n
        self.duplicates = duplicates\n
        self.wbook = openpyxl.load_workbook(self.xl.filename)\n
        self.sheet = self.wbook[self.xl.sheet_name]\n\n
    Table settings:\n
        self.table_start_col_index = 1 + self.xl.start_col\n
        self.table_start_row_index = self.xl.skip_rows + 1\n
        self.table_start = f"{openpyxl.utils.get_column_letter(self.table_start_col_index)}{self.table_start_row_index}"\n
        self.table_end_col_index = self.df.shape[1] + self.xl.start_col\n
        self.table_end_row_index = len(self.df) + self.xl.skip_rows + 1\n
        self.table_end = f"{openpyxl.utils.get_column_letter(self.table_end_col_index)}{self.table_end_row_index}"\n
    
    """
    def __init__(self, xlhandler, df_handler, duplicates) -> None:
        self.settings = Settings()
        self.xl = xlhandler
        self.df_handler = df_handler
        self.df = df_handler.df
        self.duplicates = duplicates
        self.table_start_col_index = 1 + self.xl.start_col
        self.table_start_row_index = self.xl.skip_rows + 1
        self.table_start = f"{openpyxl.utils.get_column_letter(self.table_start_col_index)}{self.table_start_row_index}"
        self.table_end_col_index = self.df.shape[1] + self.xl.start_col
        self.table_end_row_index = len(self.df) + self.xl.skip_rows + 1
        self.table_end = f"{openpyxl.utils.get_column_letter(self.table_end_col_index)}{self.table_end_row_index}"
        self.wbook = openpyxl.load_workbook(self.xl.filename)
        self.sheet = self.wbook[self.xl.sheet_name]
        self.tablename = f"Positions_REV_{self.xl.revnum}"
        self._add_table(self.tablename, self.table_start, self.table_end)
        self._table_colors()
        self._conditional_formating()
        self._add_headings()
        self._add_count()
        self._add_class_count()
        self._add_lists()
        self._adjust_sizes()
        self._alignment()
        self._add_borders(start_row= self.table_start_row_index, end_row= self.sheet.max_row + 1, start_col= self.table_start_col_index, end_col= self.table_end_col_index + 1) # Borders in Table
        self._data_validation()  
        self._hide_sheets()
        self._check_cover()
        self.wbook.active = self.sheet
        self.sheet.sheet_view.zoomScale = self.settings.data["Zoom"]["Worksheet"]
        self._save()
         

    def _conditional_formating(self):
        for colname, words in self.settings.data["Conditional"].items():
            self._conditional_col(colname, words)

    def _conditional_col(self, colname, words):  
        rownum = self.table_start_row_index  - 1
        for col in self.sheet.iter_cols(self.table_start_col_index, self.sheet.max_column):
            col_letter = openpyxl.utils.get_column_letter(col[rownum].column)
            if col[rownum].value == colname:
                for word, color in words.items():
                    colors = self.settings.data["Colors"][color]
                    fill = PatternFill()
                    if colors["Fill Color"] is not None:
                        fill = PatternFill( bgColor= colors["Fill Color"], fill_type='solid')
                    self._set_condition(col_letter, Font(color= colors["Font Color"]), fill, word)
        
    def _set_condition(self, col, text, fill, word):
        dxf = DifferentialStyle(font= text, fill=fill)
        rule = Rule(type="containsText", operator="containsText", text= word, dxf=dxf)
        rule.formula = [f'NOT(ISERROR(SEARCH("{word}",{col}{self.table_start_row_index})))']
        self.sheet.conditional_formatting.add(f'{col}{self.table_start_row_index}:{col}{self.table_end_row_index}', rule)

    def _add_table(self, tablename, start, end):
        self.table = Table(displayName= tablename, ref=f'{start}:{end}')
        style = TableStyleInfo(name="TableStyleMedium16", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        self.table.tableStyleInfo = style
        self.sheet.add_table(self.table)

    def _table_colors(self):
        color = self.settings.data["Colors"]["Scania Blue"]
        for col in range(self.table_start_col_index, self.table_end_col_index + 1):
            cell = self.sheet[f"{openpyxl.utils.get_column_letter(col)}{self.table_start_row_index}"]
            cell.font = Font(color= color["Font Color"])
            cell.fill = PatternFill(fgColor= color["Fill Color"], fill_type= 'solid')

    def _add_headings(self):
        self.sheet['B2'].value = "PART NUMBER"
        self.sheet['B2'].font = Font(bold= True)
        self.sheet['C2'].value = self.xl.partnum
        self.sheet['B3'].value = "REVISION"
        self.sheet['B3'].font = Font(bold= True)
        self.sheet['C3'].value = self.xl.revnum
        self._add_thick_outer_borders(start_row = 2, end_row= 3, start_col = 2, end_col= 3)

    def _add_count(self):
        self.sheet['D2'].value = "Audited Counts:"
        self.sheet['D2'].font = Font(bold= True)
        self.sheet['E2'].value = "Yes"
        self.sheet['E2'].alignment = Alignment(horizontal='center')
        self.sheet['F2'].value = "No"
        self.sheet['F2'].alignment = Alignment(horizontal='center')
        self.sheet['G2'].value = "Other"
        self.sheet['G2'].alignment = Alignment(horizontal='center')
        self.sheet['H2'].value = "Total"
        self.sheet['H2'].alignment = Alignment(horizontal='center')
        self.sheet['E3'].value = f'=COUNTIF({self.tablename}[Audited], "Yes")'
        self.sheet['F3'].value = f'=COUNTIF({self.tablename}[Audited], "No")'
        self.sheet['G3'].value = f'=COUNTIF({self.tablename}[Audited], "Other (comment)")'
        self.sheet['H3'].value = f'=COUNTA({self.tablename}[Position Number])'
        self._add_thick_outer_borders(start_row = 2, end_row= 3, start_col = 4, end_col= 8)

    def _add_class_count(self):
        self.sheet['J2'].value = "<C> Count:"
        self.sheet['J2'].font = Font(bold= True)
        self.sheet['J2'].alignment = Alignment(horizontal='center')
        self.sheet['J3'].value = "<M> Count:"
        self.sheet['J3'].font = Font(bold= True)
        self.sheet['J3'].alignment = Alignment(horizontal='center')
        self.sheet['J4'].value = "<L> Count:"
        self.sheet['J4'].font = Font(bold= True)
        self.sheet['J4'].alignment = Alignment(horizontal='center')
        self.sheet['K2'].value = f'=SUMPRODUCT(--EXACT("<C>",{self.tablename}[Classification]))'
        self.sheet['K3'].value = f'=SUMPRODUCT(--EXACT("<M>",{self.tablename}[Classification]))'
        self.sheet['K4'].value = f'=SUMPRODUCT(--EXACT("<L>",{self.tablename}[Classification]))'
        self._add_thick_outer_borders(start_row = 2, end_row= 4, start_col = 10, end_col= 11)


    def _add_lists(self):
        col_index = self.df_handler.col_names.index("Comment") + self.table_start_col_index
        letter_head = openpyxl.utils.get_column_letter(col_index)
        letter_list = openpyxl.utils.get_column_letter(col_index + 1)
        if len(self.duplicates) > 0:
            self.sheet[letter_head + '2'].value = "WARNING! FOUND DUPLICATES:"
            self.sheet[letter_list + '2'].value = str(self.duplicates)
            self.sheet[letter_head + '2'].font = Font(color="FF0000", bold= True)
            self.sheet[letter_list + '2'].font = Font(color="FF0000")
        self.sheet[letter_head + '3'].value = "NEW POSITION NUMBERS:"
        self.sheet[letter_head + '3'].font = Font(bold= True)
        self.sheet[letter_list + '3'].value = str(self.df_handler.new)
        self.sheet[letter_head + '4'].value = "REMOVED POSITION NUMBERS:"
        self.sheet[letter_head + '4'].font = Font(bold= True)
        self.sheet[letter_list + '4'].value = str(self.df_handler.removed)

    def _adjust_sizes(self):   
        self.sheet.column_dimensions["A"].width = 4
        self.sheet.row_dimensions[self.table_start_row_index].height = 30
        for i, col in enumerate(self.df_handler.col_names):
            col_index = i +  self.table_start_col_index
            try:
                self.sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = self.settings.data["Column Size"][col]
            except:
                self.sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = 4
          
    def _alignment(self):
        for r in range(self.table_start_row_index + 1, self.sheet.max_row + 1):
            self.sheet.cell(r, self.table_start_col_index).alignment = Alignment(horizontal='right')

    def _add_borders(self, start_row, end_row, start_col, end_col):
        side = Side(style= 'thin')
        border = Border(left= side, right= side, top= side, bottom= side)
        for row in range(start_row, end_row):
            for col in range(start_col, end_col):
                self.sheet.cell(row= row, column= col).border = border

    def _add_thick_outer_borders(self, start_row, end_row, start_col, end_col):
        thin = Side(style= 'thin')
        thick = Side(style= 'thick')
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                left, right, top, bottom = thin, thin, thin, thin
                if row == start_row:
                    top = thick
                if row == end_row:
                    bottom = thick
                if col == start_col:
                    left = thick
                if col == end_col:
                    right = thick
                self.sheet.cell(row= row, column= col).border = Border(left= left, right= right, top= top, bottom= bottom)

    def _data_validation(self):
        for key, value in self.settings.data["Data Validation"].items():
            validation_string = ", ".join(value)
            col_index = self.df.columns.get_loc(key) + self.table_start_col_index
            dv = DataValidation(type= "list", formula1= f'"{validation_string}"', allow_blank= True)
            dv.error = "Invalid Entry"
            dv.errorTitle = "Invalid Entry"
            self.sheet.add_data_validation(dv)
            for r in range(self.table_start_row_index + 1, self.sheet.max_row + 1):
                dv.add(openpyxl.utils.get_column_letter(col_index) + str(r))

            

    def _hide_sheets(self):
        for sheet in self.wbook.get_sheet_names():
            if sheet != self.xl.sheet_name:
                self.wbook[sheet].views.sheetView[0].tabSelected = False
                self.wbook[sheet].sheet_state = "hidden"

    def _check_cover(self):
        if "Cover" not in self.wbook.sheetnames:
            XlCover(self.wbook)
        XlCover(self.wbook, self.wbook['Cover']).set_part_rev(self.xl.partnum, self.xl.revnum)
        self.wbook['Cover'].sheet_state = 'visible'
        self.wbook['Cover'].sheet_view.zoomScale = self.settings.data["Zoom"]["Cover"]

    def _save(self):
        self.wbook.save(self.xl.filename)


if __name__ == "__main__":
    pass